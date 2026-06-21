# filename: python_backend/trial_balance.py
from __future__ import annotations

import json
import sys
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Font, Alignment


# ---------- sheet styling ----------

def setup_sheet_styling(target_sheet, note_text: str | None = None) -> None:
    """White background, no borders. Optional A1 note."""
    if note_text:
        target_sheet["A1"] = note_text
        target_sheet["A1"].font = Font(name="Calibri", size=11, bold=True, color="0000FF")

    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    no_border = Border()

    for row in target_sheet.iter_rows(
        min_row=1, max_row=target_sheet.max_row or 1, min_col=1, max_col=target_sheet.max_column or 1
    ):
        for cell in row:
            cell.fill = white_fill
            cell.border = no_border


# ---------- copy a validated xlsx sheet into the template, formatted ----------

def copy_xlsx_to_sheet(
    xlsx_file_path: str,
    target_workbook,
    target_sheet_name: str,
    row_offset: int = 2,
) -> tuple[str | None, str | None]:
    """Copies the FIRST sheet of `xlsx_file_path` (our validated files only ever
    contain one sheet) into target_workbook as target_sheet_name, pasted at A3.
    Returns (actual_source_sheet_name, error_message)."""
    try:
        source_wb = load_workbook(xlsx_file_path, data_only=False)
    except Exception as exc:
        return None, f"Unable to open validated file: {exc}"

    if not source_wb.sheetnames:
        return None, "Validated file has no sheets."

    source_sheet_name = source_wb.sheetnames[0]
    source_sheet = source_wb[source_sheet_name]
    target_sheet = target_workbook.create_sheet(title=target_sheet_name)

    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row + row_offset, column=cell.column)
            target_cell.value = cell.value

            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = Border()
                target_cell.fill = PatternFill(fill_type=None)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width

    for row_num, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num + row_offset].height = row_dim.height

    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(
            start_row=merged_range.min_row + row_offset,
            start_column=merged_range.min_col,
            end_row=merged_range.max_row + row_offset,
            end_column=merged_range.max_col,
        )

    setup_sheet_styling(
        target_sheet,
        note_text=(
            "D&T downloaded the below statement as part of the Audit Package along "
            "with the Limited Scope certification using our auditor's access."
        ),
    )

    return source_sheet_name, None


# ---------- read a validated xlsx sheet into a DataFrame (header = row 1) ----------

def read_validated_df(xlsx_file_path: str, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(
        xlsx_file_path,
        sheet_name=sheet_name,
        engine="openpyxl",
        header=0,
        dtype=object,
    )


# ---------- hierarchical summary ----------

def summarize_hierarchical_df(
    df: pd.DataFrame,
    index_col: str,
    sum_col: str,
    group_headers: list[str],
    section_groups: list[str],
    roll_groups: list[str],
    combine_groups: dict[str, list[str]],
    target_workbook,
    sheet_name: str,
) -> dict[str, float]:
    df = df.copy()
    df[sum_col] = (
        df[sum_col]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.strip()
        .replace({"": "0", "-": "0", "N/A": "0", "n/a": "0", "nan": "0", "None": "0"})
    )
    df[sum_col] = pd.to_numeric(df[sum_col], errors="coerce").fillna(0.0)

    header_lookup = {h.strip().lower(): h for h in group_headers}
    section_lookup = {s.strip().lower() for s in section_groups}
    roll_lookup = {r.strip().lower() for r in roll_groups}

    combine_lookup: dict[str, str] = {}
    for output_name, input_headers in combine_groups.items():
        for h in input_headers:
            combine_lookup[h.strip().lower()] = output_name

    raw_results: dict[str, float] = {}
    current_section: str | None = None

    for _, row in df.iterrows():
        label = row[index_col]

        if pd.isna(label) or str(label).strip() == "" or str(label).strip().lower() == "nan":
            current_section = None
            continue

        label_str = str(label).strip()
        label_lower = label_str.lower()
        num_val = float(row[sum_col])

        if label_lower in header_lookup:
            canonical = header_lookup[label_lower]
            canon_lower = canonical.lower()

            if canon_lower in section_lookup:
                current_section = canonical
                raw_results.setdefault(canonical, 0.0)
                if canon_lower in roll_lookup:
                    raw_results.setdefault(f"{canonical} (Rollover)", 0.0)
                continue

            current_section = None
            if canon_lower in combine_lookup:
                out_name = combine_lookup[canon_lower]
                raw_results[out_name] = raw_results.get(out_name, 0.0) + num_val
            else:
                raw_results[canonical] = raw_results.get(canonical, 0.0) + num_val
            continue

        if current_section is not None:
            sec_lower = current_section.lower()
            if sec_lower in roll_lookup and "roll" in label_lower:
                raw_results[f"{current_section} (Rollover)"] += num_val
            else:
                raw_results[current_section] += num_val

    consumed = {h.strip().lower() for headers in combine_groups.values() for h in headers}
    results = {k: v for k, v in raw_results.items() if k.lower() not in consumed}

    ws = target_workbook.create_sheet(title=sheet_name)
    bold_font = Font(name="Calibri", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center = Alignment(horizontal="center")

    for col_idx, heading in enumerate(["Description", "Total Plan Activity per FS"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=heading)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    for row_num, (group_name, total) in enumerate(results.items(), start=2):
        ws.cell(row=row_num, column=1, value=group_name)
        total_cell = ws.cell(row=row_num, column=2, value=round(total, 2))
        total_cell.number_format = "#,##0.00"

    return results


# ---------- orchestration ----------

def build_trial_balance(payload: dict[str, Any]) -> dict[str, Any]:
    log: list[str] = []
    template_path = payload.get("template_path")
    output_path = payload.get("output_path")
    reporting_date = payload.get("reporting_date", "")
    rows = payload.get("rows", [])
    summary_cfg = payload.get("summary")

    if not template_path or not Path(template_path).exists():
        return {"success": False, "message": "Template workbook not found.", "log": log}

    if not output_path:
        return {"success": False, "message": "No output path provided.", "log": log}

    try:
        template_wb = load_workbook(template_path)
    except Exception as exc:
        return {"success": False, "message": f"Unable to open template: {exc}", "log": log}

    if reporting_date and "1. Procedures" in template_wb.sheetnames:
        template_wb["1. Procedures"]["A2"] = reporting_date

    copied_sheets: list[dict[str, Any]] = []
    summary_df = None
    summary_index_col = None
    summary_sum_col = None

    for row in rows:
        validated_path = row.get("validated_path")
        target_sheet_name = row.get("target_sheet_name")
        row_id = row.get("row_id")

        if not validated_path or not target_sheet_name:
            log.append(f"Skipped row {row_id}: missing validated_path or target_sheet_name.")
            continue

        if not Path(validated_path).exists():
            log.append(f"Skipped row {row_id}: validated file not found at {validated_path}.")
            continue

        source_sheet_name, error = copy_xlsx_to_sheet(validated_path, template_wb, target_sheet_name)
        if error:
            log.append(f"Row {row_id}: {error}")
            continue

        log.append(f"Row {row_id}: copied '{source_sheet_name}' -> '{target_sheet_name}'.")
        copied_sheets.append({"rowId": row_id, "targetSheetName": target_sheet_name})

        if summary_cfg and row.get("use_for_summary"):
            try:
                summary_df = read_validated_df(validated_path, source_sheet_name)
                summary_index_col = row.get("index_col")
                summary_sum_col = row.get("sum_col")
            except Exception as exc:
                log.append(f"Row {row_id}: unable to read for summary: {exc}")

    summary_result = None
    if summary_cfg and summary_df is not None and summary_index_col and summary_sum_col:
        try:
            summary_result = summarize_hierarchical_df(
                df=summary_df,
                index_col=summary_index_col,
                sum_col=summary_sum_col,
                group_headers=summary_cfg.get("group_headers", []),
                section_groups=summary_cfg.get("section_groups", []),
                roll_groups=summary_cfg.get("roll_groups", []),
                combine_groups=summary_cfg.get("combine_groups", {}),
                target_workbook=template_wb,
                sheet_name=summary_cfg.get("sheet_name", "Summary"),
            )
            log.append(f"Summary sheet '{summary_cfg.get('sheet_name', 'Summary')}' written.")
        except Exception as exc:
            log.append(f"Summary generation failed: {exc}")

    try:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        template_wb.save(output_path)
    except Exception as exc:
        return {"success": False, "message": f"Unable to save output: {exc}", "log": log}

    return {
        "success": True,
        "outputPath": str(Path(output_path).resolve()),
        "copiedSheets": copied_sheets,
        "summary": summary_result,
        "log": log,
    }


def main() -> None:
    try:
        payload = json.load(sys.stdin)
        result = build_trial_balance(payload)
        print(json.dumps(result))
        sys.exit(0 if result.get("success") else 1)
    except Exception as exc:
        print(json.dumps({"success": False, "message": str(exc), "log": []}))
        sys.exit(1)


if __name__ == "__main__":
    main()