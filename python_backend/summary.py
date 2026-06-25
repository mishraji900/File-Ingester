from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from copy import copy
import xlrd
from xlrd import open_workbook
# from xlrd.formatting import import XFStyle
import pandas as pd
import datetime
import csv
import os

##### Summarizing DFs

# def summarize_hierarchical_df(
#     df: pd.DataFrame,
#     index_col,                  # column name (str) or integer position for labels
#     sum_col,                    # column name (str) or integer position for values
#     group_headers: list[str],   # all header labels e.g. ["Contri/Emp", "contri/Emp2", "adj+", "adj-", ...]
#     roll_groups: list[str],     # groups needing a roll split row e.g. ["contri/Emp2"]
#     combine_groups: dict[str, list[str]], # merged output rows e.g. {"net adj": ["adj+", "adj-"]}
#     template_wb,                # openpyxl Workbook to write into
#     sheet_name: str,            # name for the new summary sheet
# ) -> dict:
#     """
#     Scans a hierarchically structured DataFrame and produces group totals.
#     
#     - Groups are delimited by header rows found in index_col.
#     - Blank rows are skipped.
#     - sum_col is coerced to float at the start to handle object dtype.
#     - roll_groups: rows whose label contains 'roll' go to '{group} (Roll)' row;
#       all other rows go to the normal group row.
#     - combine_groups: multiple header groups whose totals are summed into one
#       output row (e.g. adj+ and adj- combined into "net adj").
#       
#     Returns a dict: { output_row_name: {"total": float} }
#     """
#     
#     # — Coerce sum_col from object dtype to float upfront ——————————————
#     col_name = sum_col if isinstance(sum_col, str) else df.columns[sum_col]
#     idx_name = index_col if isinstance(index_col, str) else df.columns[index_col]
#     
#     df = df.copy()
#     df[col_name] = (
#         df[col_name]
#         .astype(str)
#         .str.replace(",", "", regex=False)
#         .str.replace("$", "", regex=False)
#         .str.strip()
#         .replace({"": "0", "-": "0", "N/A": "0", "n/a": "0", "nan": "0", "None": "0"})
#     )
#     df[col_name] = pd.to_numeric(df[col_name], errors="coerce").fillna(0.0)
#     
#     # — Build lookup tables ——————————————
#     # case-insensitive header -> original cased label
#     header_lookup = {h.strip().lower(): h for h in group_headers}
#     
#     # case-insensitive set for roll groups
#     roll_lookup = {r.strip().lower() for r in roll_groups}
#     
#     # reverse map: individual header lower -> combined output name
#     # e.g. {"adj+": "net adj", "adj-": "net adj"}
#     combine_lookup: dict[str, str] = {}
#     for output_name, input_headers in combine_groups.items():
#         for h in input_headers:
#             combine_lookup[h.strip().lower()] = output_name
#             
#     # — Accumulation dict - internal, keyed by original header ——————————————
#     raw_results: dict[str, float] = {}
#     current_group: str | None = None
#     
#     for _, row in df.iterrows():
#         label = row[idx_name]
#         
#         # Skip blank / NaN rows
#         if pd.isna(label) or str(label).strip() == "" or str(label).strip().lower() == "nan":
#             continue
#             
#         label_str = str(label).strip()
#         label_lower = label_str.lower()
#         
#         # — Group header detected ——————————————
#         if label_lower in header_lookup:
#             current_group = header_lookup[label_lower]
#             current_lower = current_group.lower()
#             
#             # Initialize normal row
#             if current_group not in raw_results:
#                 raw_results[current_group] = 0.0
#                 
#             # Initialize roll row if needed
#             if current_lower in roll_lookup:
#                 roll_key = f"{current_group} (Roll)"
#                 if roll_key not in raw_results:
#                     raw_results[roll_key] = 0.0
#                     
#             # Initialize combined output row if needed
#             if current_lower in combine_lookup:
#                 out_name = combine_lookup[current_lower]
#                 if out_name not in raw_results:
#                     raw_results[out_name] = 0.0
#                     
#             continue  # header row carries no numeric value
#             
#         # — Data row ——————————————
#         if current_group is None:
#             continue
#             
#         num_val = float(row[col_name])
#         current_lower = current_group.lower()
#         
#         # Case 1: roll group - split on 'roll' substring
#         if current_lower in roll_lookup:
#             if "roll" in label_lower:
#                 raw_results[f"{current_group} (Roll)"] += num_val
#             else:
#                 raw_results[current_group] += num_val
#                 
#         # Case 2: combined group (e.g. adj+, adj-) - accumulate into combined key
#         elif current_lower in combine_lookup:
#             out_name = combine_lookup[current_lower]
#             raw_results[out_name] += num_val
#             
#         # Case 3: regular group
#         else:
#             raw_results[current_group] += num_val
#             
#     # — Build final ordered results (exclude raw headers consumed by combine) ——————————————
#     consumed = {h.strip().lower() for headers in combine_groups.values() for h in headers}
#     results: dict[str, float] = {
#         k: v for k, v in raw_results.items()
#         if k.lower() not in consumed              # drop adj+/adj- raw keys
#         or k.lower() not in combine_lookup       # keep anything not in combine
#     }
#     
#     # — Write summary to template_wb ——————————————
#     ws = template_wb.create_sheet(title=sheet_name)
#     
#     bold_font   = Font(name="Calibri", bold=True, size=11)
#     header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
#     center      = Alignment(horizontal="center")
#     no_border   = Border(
#         left=Side(style=None), right=Side(style=None),
#         top=Side(style=None), bottom=Side(style=None),
#     )
#     
#     # Column headers
#     for col_idx, heading in enumerate(["Group", "Total"], start=1):
#         cell = ws.cell(row=1, column=col_idx, value=heading)
#         cell.font      = bold_font
#         cell.fill      = header_fill
#         cell.alignment = center
#         cell.border    = no_border
#         
#     ws.column_dimensions["A"].width = 30
#     ws.column_dimensions["B"].width = 18
#     
#     # Data rows
#     for row_num, (group_name, total) in enumerate(results.items(), start=2):
#         ws.cell(row=row_num, column=1, value=group_name).border = no_border
#         total_cell = ws.cell(row=row_num, column=2, value=round(total, 2))
#         total_cell.number_format = "#,##0.00"
#         total_cell.border        = no_border
#         
#     print(f"✓ Summary written to sheet '{sheet_name}' – {len(results)} row(s) written")
#     return results


def summarize_hierarchical_df(
    df: pd.DataFrame,
    index_col,
    sum_col,
    group_headers: list[str],
    section_groups: list[str],    # 💡 NEW: only these act as section headers with sub-rows
    roll_groups: list[str],
    combine_groups: dict[str, list[str]],
    template_wb,
    sheet_name: str,
) -> dict:
    """
    - section_groups: labels that act as section headers; rows BELOW them are summed.
      e.g. ["Contributions/Employer", "Contributions/Employee"]
    - group_headers: all labels to track (sections + direct rows).
    - roll_groups: section groups where sub-rows with 'roll' split into a separate row.
    - combine_groups: direct-row groups merged into one output row.
      e.g. {"Net Adjustments": ["Adjustment (+)", "Adjustment (-)"]}
    """
    
    # — Coerce sum_col to float ——————————————
    col_name = sum_col if isinstance(sum_col, str) else df.columns[sum_col]
    idx_name = index_col if isinstance(index_col, str) else df.columns[index_col]
    
    df = df.copy()
    df[col_name] = (
        df[col_name]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.strip()
        .replace({"": "0", "-": "0", "N/A": "0", "n/a": "0", "nan": "0", "None": "0"})
    )
    df[col_name] = pd.to_numeric(df[col_name], errors="coerce").fillna(0.0)
    
    # — Lookup tables ——————————————
    header_lookup  = {h.strip().lower(): h for h in group_headers}
    section_lookup = {s.strip().lower() for s in section_groups}   # section headers only
    roll_lookup    = {r.strip().lower() for r in roll_groups}
    
    # reverse map: direct-row label lower -> combined output name
    combine_lookup: dict[str, str] = {}
    for output_name, input_headers in combine_groups.items():
        for h in input_headers:
            combine_lookup[h.strip().lower()] = output_name
            
    # — Accumulation ——————————————
    raw_results: dict[str, float] = {}
    current_section: str | None = None
    
    for _, row in df.iterrows():
        label = row[idx_name]
        
        if pd.isna(label) or str(label).strip() == "" or str(label).strip().lower() == "nan":
            current_section = None  # blank row resets section context
            continue
            
        label_str = str(label).strip()
        label_lower = label_str.lower()
        num_val = float(row[col_name])
        
        # — CASE A: label is a known group header ——————————————
        if label_lower in header_lookup:
            canonical = header_lookup[label_lower]
            canon_lower = canonical.lower()
            
            # Sub-case A1: SECTION header – sub-rows follow, don't sum this row
            if canon_lower in section_lookup:
                current_section = canonical
                if canonical not in raw_results:
                    raw_results[canonical] = 0.0
                if canon_lower in roll_lookup:
                    roll_key = f"{canonical} (Rollover)"
                    if roll_key not in raw_results:
                        raw_results[roll_key] = 0.0
                continue  # move to sub-rows
                
            # Sub-case A2: DIRECT row – sum its value immediately
            else:
                current_section = None  # direct rows don't open a section
                
                # Goes into a combined output row (e.g. net adj)
                if canon_lower in combine_lookup:
                    out_name = combine_lookup[canon_lower]
                    raw_results[out_name] = raw_results.get(out_name, 0.0) + num_val
                    
                # Standalone direct row
                else:
                    raw_results[canonical] = raw_results.get(canonical, 0.0) + num_val
                    
            continue
            
        # — CASE B: sub-row inside a section ——————————————
        if current_section is not None:
            sec_lower = current_section.lower()
            if sec_lower in roll_lookup and "roll" in label_lower:
                raw_results[f"{current_section} (Rollover)"] += num_val
            else:
                raw_results[current_section] += num_val
                
        # rows that match nothing and are outside a section are ignored
        
    # — Filter out raw combined-input keys from output ——————————————
    consumed = {h.strip().lower() for headers in combine_groups.values() for h in headers}
    results: dict[str, float] = {
        k: v for k, v in raw_results.items()
        if k.lower() not in consumed
    }
    
    # — Write to Excel ——————————————
    ws = template_wb.create_sheet(title=sheet_name)
    
    bold_font   = Font(name="Calibri", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center      = Alignment(horizontal="center")
    no_border   = Border(
        left=Side(style=None), right=Side(style=None),
        top=Side(style=None), bottom=Side(style=None),
    )
    
    for col_idx, heading in enumerate(["Description", "Total Plan Activity per FS"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=heading)
        cell.font      = bold_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = no_border
        
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    
    for row_num, (group_name, total) in enumerate(results.items(), start=2):
        ws.cell(row=row_num, column=1, value=group_name).border = no_border
        total_cell = ws.cell(row=row_num, column=2, value=round(total, 2))
        total_cell.number_format = "#,##0.00"
        total_cell.border        = no_border
        
    print(f"✓ Summary written to sheet '{sheet_name}' – {len(results)} row(s) written")
    return results


##### DF Readers

def read_sheet_to_df(file_path: str, sheet_name: str | None = None, header_row: int | None = None) -> pd.DataFrame | None:
    """
    Unified dispatcher – reads a sheet (or CSV) into a pandas DataFrame.
    
    - Row 1 is always treated as DATA, never as a header.
    - Column names are auto-assigned as integers (0, 1, 2, ...).
    - Handles .xlsx, .xlsm, .xls, and .csv.
    
    Args:
        file_path:  Path to the source file.
        sheet_name: Sheet name to read (ignored for CSV files).
        
    Returns:
        A pandas DataFrame, or None on failure.
    """
    ext = os.path.splitext(file_path)[1].lower()
    # Convert 1-based -> 0-based for pandas; keep None as-is
    header_idx = (header_row - 1) if header_row is not None else None
    try:
        if ext in (".xlsx", ".xlsm"):
            return _read_xlsx_sheet_to_df(file_path, sheet_name, header_idx)
        elif ext == ".xls":
            return _read_xls_sheet_to_df(file_path, sheet_name, header_idx)
        elif ext == ".csv":
            return _read_csv_to_df(file_path, header_idx)
        else:
            print(f"⚠️ Unsupported extension for DataFrame read: {ext}")
            return None
    except Exception as e:
        print(f"❌ Error reading '{sheet_name}' from {file_path} into DataFrame: {e}")
        return None


def _read_xlsx_sheet_to_df(file_path: str, sheet_name: str, header_idx: int | None) -> pd.DataFrame:
    """Read one sheet from an .xlsx/.xlsm into a DataFrame (no header row)."""
    return pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        engine="openpyxl",
        header=header_idx,
        dtype=object,            # preserve mixed types without coercion
    )


def _read_xls_sheet_to_df(file_path: str, sheet_name: str, header_idx: int | None) -> pd.DataFrame:
    xls_wb = open_workbook(file_path, formatting_info=False)
    
    if sheet_name not in xls_wb.sheet_names():
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {xls_wb.sheet_names()}")
        
    ws = xls_wb.sheet_by_name(sheet_name)
    data = []
    
    for row_idx in range(ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)
            ctype, value = cell.ctype, cell.value
            
            if ctype == xlrd.XL_CELL_DATE:
                try:
                    parts = xlrd.xldate_as_tuple(value, xls_wb.datemode)
                    value = (
                        datetime.time(parts[3], parts[4], int(parts[5]))
                        if parts[0] == 0
                        else datetime.datetime(*parts[:3], *parts[3:5], int(parts[5]))
                    )
                except Exception:
                    pass
            elif ctype == xlrd.XL_CELL_BOOLEAN:
                value = bool(value)
            elif ctype == xlrd.XL_CELL_ERROR:
                value = None
                
            row.append(value)
        data.append(row)
        
    df = pd.DataFrame(data)
    
    # ADD: promote the chosen row to column names, then drop it from data
    if header_idx is not None:
        df.columns = df.iloc[header_idx]
        df = df.drop(index=header_idx).reset_index(drop=True)
        
    return df


def _read_csv_to_df(file_path: str, header_idx: int | None, encoding: str = "utf-8") -> pd.DataFrame:
    df = pd.read_csv(file_path, header=header_idx, encoding=encoding, dtype=object) # was hardcoded None
    
    def _coerce(val):
        if pd.isna(val):
            return val
        try:
            return int(val) if "." not in str(val) else float(val)
        except (ValueError, TypeError):
            return val
            
    return df.apply(lambda col: col.map(_coerce))


def setup_sheet_styling(target_sheet):
    """
    Apply white background and remove borders from the entire sheet.
    Set A1 to statement.
    """
    # Set A1 text
    target_sheet['A1'] = "D&T downloaded the below statement as part of the Audit Package along with the Limited Scope certification at wp 28190.1 using our auditor's access."
    target_sheet['A1'].font = Font(name='Calibri', size=11, bold=True, color='0000FF')
    
    # Solid white fill (this is key - not fill_type=None)
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # No borders
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    
    # Apply to all cells that will be used (estimate a reasonable range)
    for row in target_sheet.iter_rows(min_row=1, max_row=9999, min_col=1, max_col=100):
        for cell in row:
            cell.fill = white_fill
            cell.border = no_border


def _shift_cf_range(range_str: str, row_offset: int) -> str | None:
    """Shift a conditional formatting range string down by row_offset rows."""
    from openpyxl.utils.cell import range_boundaries, get_column_letter
    try:
        parts = range_str.split(":")
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        new_min = f"{get_column_letter(min_col)}{min_row + row_offset}"
        new_max = f"{get_column_letter(max_col)}{max_row + row_offset}"
        return f"{new_min}:{new_max}"
    except Exception:
        return None


def _shift_sqref(sqref, row_offset: int):
    """Shift a data-validation sqref (CellRange or string) down by row_offset rows."""
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    
    result_parts = []
    for part in str(sqref).split():
        try:
            min_col, min_row, max_col, max_row = range_boundaries(part)
            new_start = f"{get_column_letter(min_col)}{min_row + row_offset}"
            new_end = f"{get_column_letter(max_col)}{max_row + row_offset}"
            result_parts.append(f"{new_start}:{new_end}")
        except Exception:
            result_parts.append(part)  # leave as-is if unparseable
    return " ".join(result_parts)


# def copy_xlsx_to_sheet(
#     xlsx_file_path: str,
#     sheet_name: str,
#     target_workbook,
#     target_sheet_name: str,
#     row_offset: int = 2,
#     preserve_borders: bool = False,
#     preserve_fills: bool = True,
# ) -> object | None:
#     """
#     Copy a sheet from a source .xlsx to a target workbook with full formatting.
#     
#     Args:
#         xlsx_file_path:    Path to the source .xlsx file.
#         sheet_name:        Name of the sheet to copy from the source.
#         target_workbook:   The openpyxl Workbook object to copy into.
#         target_sheet_name: Name for the new sheet in the target workbook.
#         row_offset:        Number of rows to shift content down (default=2 -> paste at A3).
#         preserve_borders:  If True, copies source borders; if False, strips all borders.
#         preserve_fills:    If True, copies source fills; if False, applies no fill (transparent).
#         
#     Returns:
#         The created target worksheet, or None on failure.
#     """
#     try:
#         source_workbook = load_workbook(xlsx_file_path, data_only=False)
#         
#         if sheet_name not in source_workbook.sheetnames:
#             print(f"⚠️ Warning: Sheet '{sheet_name}' not found in {xlsx_file_path}")
#             return None
#             
#         source_sheet = source_workbook[sheet_name]
#         target_sheet = target_workbook.create_sheet(title=target_sheet_name)
        
        # # — 1. Cell values + formatting ——————————————
        # for row in source_sheet.iter_rows():
        #     for cell in row:
        #         new_row = cell.row + row_offset
        #         target_cell = target_sheet.cell(row=new_row, column=cell.column)
        #         
        #         # Value (preserve formula strings as-is)
        #         target_cell.value = cell.value
        #         
        #         if cell.has_style:
        #             target_cell.font = copy(cell.font)
        #             target_cell.alignment = copy(cell.alignment)
        #             target_cell.protection = copy(cell.protection)
        #             
        #             # number_format is a plain string – direct assignment is correct
        #             target_cell.number_format = cell.number_format
        #             
        #             # Border: preserve or strip based on flag
        #             target_cell.border = copy(cell.border) if preserve_borders else Border()
        #             
        #             # Fill: preserve or strip based on flag
        #             if preserve_fills:
        #                 target_cell.fill = copy(cell.fill)
        #             else:
        #                 target_cell.fill = PatternFill(fill_type=None)
                        
        # # — 2. Column dimensions ——————————————
        # for col_letter, col_dim in source_sheet.column_dimensions.items():
        #     target_col_dim = target_sheet.column_dimensions[col_letter]
        #     target_col_dim.width = col_dim.width
        #     target_col_dim.hidden = col_dim.hidden
        #     target_col_dim.outlineLevel = col_dim.outlineLevel
        #     
        # # — 3. Row dimensions (shifted by offset) ——————————————
        # for row_num, row_dim in source_sheet.row_dimensions.items():
        #     target_row_dim = target_sheet.row_dimensions[row_num + row_offset]
        #     target_row_dim.height = row_dim.height
        #     target_row_dim.hidden = row_dim.hidden
        #     target_row_dim.outlineLevel = row_dim.outlineLevel
            
        # # — 4. Merged cells (shifted by offset) ——————————————
        # for merged_range in source_sheet.merged_cells.ranges:
        #     target_sheet.merge_cells(
        #         start_row=merged_range.min_row + row_offset,
        #         start_column=merged_range.min_col,
        #         end_row=merged_range.max_row + row_offset,
        #         end_column=merged_range.max_col,
        #     )
            
        # # — 5. Conditional formatting (shifted by offset) ——————————————
        # for cf_range, cf_rules in source_sheet.conditional_formatting._cf_rules.items():
        #     for rule in cf_rules:
        #         # Shift each cell range in the CF address
        #         shifted_ranges = _shift_cf_range(str(cf_range), row_offset)
        #         if shifted_ranges:
        #             target_sheet.conditional_formatting.add(shifted_ranges, copy(rule))
                    
        # # — 6. Data validations (shifted by offset) ——————————————
        # for dv in source_sheet.data_validations.dataValidation:
        #     new_dv = copy(dv)
        #     new_dv.sqref = _shift_sqref(dv.sqref, row_offset)
        #     target_sheet.add_data_validation(new_dv)
            
        # # — 7. Sheet-level styling (called ONCE, after everything is set) ——————————————
        # setup_sheet_styling(target_sheet)
        
        # return target_sheet
        
    # except Exception as e:
    #     print(f"❌ Error copying sheet: {e}")
    #     raise  # Re-raise so the caller can handle or log properly


#### Functions to copy raw to template

def copy_xlsx_to_sheet(xlsx_file_path, sheet_name, target_workbook, target_sheet_name):
    """
    Copy a sheet from source to target workbook with formatting, starting at A3.
    """
    try:
        xlsx_workbook = load_workbook(xlsx_file_path, data_only=False)
        # Get the sheet by name
        if sheet_name not in xlsx_workbook.sheetnames:
            print(f"⚠️ Warning: Sheet '{sheet_name}' not found in '{xlsx_file_path}'")
            return None
            
        source_sheet = xlsx_workbook[sheet_name]
        
        # Create new sheet in target workbook
        target_sheet = target_workbook.create_sheet(title=target_sheet_name)
        
        # # Setup sheet styling (white background, no borders, A1 text)
        # setup_sheet_styling(target_sheet)
        
        # Offset for pasting at A3 (2 rows down)
        row_offset = 2
        
        # Copy cell values and formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                # Calculate new position with offset
                new_row = cell.row + row_offset
                new_col = cell.column
                target_cell = target_sheet.cell(row=new_row, column=new_col)
                # Copy value
                target_cell.value = cell.value

                # Copy formatting but override borders
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = Border()  # No borders
                    target_cell.fill = PatternFill(fill_type=None)  # White background
                    target_cell.number_format = copy(cell.number_format)
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)

        # Copy column dimensions
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width

        # Copy row dimensions (with offset)
        for row_num in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row_num + row_offset].height = source_sheet.row_dimensions[row_num].height

        # Copy merged cells (with offset)
        for merged_cell in source_sheet.merged_cells.ranges:
            min_col = merged_cell.min_col
            min_row = merged_cell.min_row + row_offset
            max_col = merged_cell.max_col
            max_row = merged_cell.max_row + row_offset
            target_sheet.merge_cells(start_row=min_row, start_column=min_col,
                                     end_row=max_row, end_column=max_col)

        # Setup sheet styling (white background, no borders, A1 text)
        setup_sheet_styling(target_sheet)

        return target_sheet

    except Exception as e:
        print(f"❌ Error reading .xlsx file: {e}")
        return None


def copy_xls_to_sheet(xls_file_path, sheet_name, target_workbook, target_sheet_name):
    """
    Read an .xls file and copy specified sheet to target workbook with formatting preserved.
    Data starts at A3 in the target sheet.
    """
    try:
        # Open the .xls file with formatting info
        xls_workbook = open_workbook(xls_file_path, formatting_info=True)

        # Get the sheet by name
        if sheet_name not in xls_workbook.sheet_names():
            print(f"⚠️ Warning: Sheet '{sheet_name}' not found in {xls_file_path}")
            print(f"Available sheets: {xls_workbook.sheet_names()}")
            return None

        source_sheet = xls_workbook.sheet_by_name(sheet_name)

        # Create new sheet in target workbook
        target_sheet = target_workbook.create_sheet(title=target_sheet_name)
        # setup_sheet_styling(target_sheet)

        # Offset for pasting at A3 (2 rows down)
        row_offset = 2

        # Get formatting book
        formatting_book = xls_workbook.formatting_info

        # Copy cell values and formatting
        for row_idx in range(source_sheet.nrows):
            for col_idx in range(source_sheet.ncols):
                # Get source cell
                cell_value = source_sheet.cell_value(row_idx, col_idx)
                cell_type = source_sheet.cell_type(row_idx, col_idx)

                # Get target cell with offset
                target_cell = target_sheet.cell(row=row_idx + 1 + row_offset, column=col_idx + 1)
                target_cell.value = cell_value

                # Get formatting from source
                try:
                    xf_index = source_sheet.cell_xf_index(row_idx, col_idx)
                    xf = xls_workbook.xf_list[xf_index]

                    # Apply number format
                    format_str = xf.format_key
                    # Convert XLS format to XLSX format
                    target_cell.number_format = target_cell.number_format

                    # Get cell format
                    cell_xf = xls_workbook.xf_list[xf_index]

                    # Apply font formatting
                    font_index = cell_xf.font_index
                    font = xls_workbook.font_list[font_index]
                    target_cell.font = Font(
                        name=font.name,
                        size=font.height / 20,  # Convert twips to points
                        bold=font.weight > 400,
                        italic=font.italic,
                        color=None  # Keep default color
                    )

                    # Apply alignment
                    alignment = cell_xf.alignment
                    horiz_align_map = {0: 'general', 1: 'left', 2: 'center', 3: 'right', 4: 'fill', 5: 'justify'}
                    vert_align_map = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify'}
                    
                    target_cell.alignment = Alignment(
                        horizontal=horiz_align_map.get(alignment.hor_align, 'general'),
                        vertical=vert_align_map.get(alignment.vert_align, 'bottom'),
                        wrap_text=bool(alignment.text_wrapped)
                    )

                except Exception as format_error:
                    # If formatting fails, continue with just the value
                    pass

                # Override borders and fill for consistency
                target_cell.border = Border()
                target_cell.fill = PatternFill(fill_type=None)

        # Copy column widths
        for col_idx in range(source_sheet.ncols):
            try:
                col_info = source_sheet.colinfo_map.get(col_idx)
                if col_info:
                    # Convert width from XLS units to Excel units
                    width = col_info.width / 256
                    col_letter = get_column_letter(col_idx + 1)
                    target_sheet.column_dimensions[col_letter].width = width
            except:
                pass

        # Copy row heights
        for row_idx in range(source_sheet.nrows):
            try:
                row_info = source_sheet.rowinfo_map.get(row_idx)
                if row_info:
                    # Convert height from twips to points
                    height = row_info.height / 20
                    target_sheet.row_dimensions[row_idx + 1 + row_offset].height = height
            except:
                pass

        print(f"✓ Sheet '{sheet_name}' from .xls file copied as '{target_sheet_name}' with formatting")
        setup_sheet_styling(target_sheet)

        return target_sheet

    except Exception as e:
        print(f"❌ Error reading .xls file: {e}")
        import traceback
        traceback.print_exc()
        return None


def copy_csv_to_sheet(csv_file_path, sheet_name, target_workbook, target_sheet_name, encoding='utf-8'):
    """
    Read a CSV file and copy its contents to a new sheet in the target workbook.
    Data starts at A3 in the target sheet.
    """
    try:
        # Create new sheet in target workbook
        target_sheet = target_workbook.create_sheet(title=target_sheet_name)
        
        # Setup sheet styling (white background, no borders, A1 text)
        setup_sheet_styling(target_sheet)
        
        # Offset for pasting at A3 (2 rows down)
        row_offset = 2
        
        # Read and copy CSV data
        with open(csv_file_path, 'r', encoding=encoding, newline='') as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=',', quotechar='"')
            
            for row_idx, row_data in enumerate(csv_reader, start=1):
                # Apply row offset
                target_row = row_idx + row_offset
                for col_idx, cell_value in enumerate(row_data, start=1):
                    target_cell = target_sheet.cell(row=target_row, column=col_idx)
                    
                    # Try to convert to number if possible
                    try:
                        if '.' in cell_value:
                            target_cell.value = float(cell_value)
                        else:
                            target_cell.value = int(cell_value)
                    except (ValueError, TypeError):
                        target_cell.value = cell_value
                        
                    # Ensure no borders and white background
                    target_cell.border = Border()
                    target_cell.fill = PatternFill(fill_type=None)
                    
        print(f"✓ CSV file copied as '{target_sheet_name}'")
        setup_sheet_styling(target_sheet)
        
        return target_sheet
        
    except Exception as e:
        print(f"❌ Error reading CSV file: {e}")
        return None


def main():
    # File paths - SS: can take from electron UI
    file_SNTA_path = "converted/Summary of Net Trust Assets 12-31-24.xlsx"
    file_SONTA_path = "converted/Summary of Plan Operations 12-31-24.xlsx"
    file_SOP_path = "converted/SOP"  # SS: can take from UI
    template_path = "../runtime/Workpaper Template/28180 Trial Balance.xlsx"
    output_path = "../runtime/Outputs/28180 Trial Balance.xlsx"
    file_extension = ".xlsx"  # SS: can change from electron UI
    ### SS: can take header row number from UI
    header_row_sonta = 6
    header_row_loans = 5
    
    try:
        # Load the workbooks
        print("Program in execution mode...")
        template_wb = load_workbook(template_path)
        date_sheet = template_wb["1. Procedures"]
        date_sheet['A2'] = "12/31/2024"
        print(f"Current working directory: {os.getcwd()}")
        print(f"os.path.exists('converted'): {os.path.exists('converted')}")
        
        # Reading files on extensions
        if file_extension == ".xlsx":
            # Copy from .xlsx file (if file exists)
            if os.path.exists(file_SNTA_path):
                print(f"Copying from .xlsx file: {file_SNTA_path}...")
                copy_xlsx_to_sheet(file_SNTA_path, file_SNTA_path, template_wb, "2. Summary of Net Trust Asset")
                # SOP0 file - Sheet TOTALS
            if os.path.exists(file_SOPO_path):
                print(f"Copying from .xlsx file: {file_SOPO_path}...")
                copy_xlsx_to_sheet(file_SOPO_path, file_SOPO_path, template_wb, "3. Summary of Plan Ops")
                
            if os.path.exists(file_SOP_path):
                print(f"Copying from .xlsx file: {file_SOP_path}...")
                copy_xlsx_to_sheet(file_SOP_path, file_SOPO_path, template_wb, "4. Loan")
                
            snta_df = read_sheet_to_df(file_SNTA_path, file_SNTA_path, header_idx)
            totals_df = read_sheet_to_df(file_SOPO_path, file_SOPO_path, header_row_totals)
            
        elif file_extension == ".xls":
            # Copy from .xls file (if file exists)
            if os.path.exists(file_SNTA_path):
                print(f"Copying from .xls file: {file_SNTA_path}...")
                copy_xls_to_sheet(file_SNTA_path, file_SNTA_path, template_wb, "2. Summary of Net Trust Asset")
                # SOP0 file - Sheet TOTALS
            if os.path.exists(file_SOPO_path):
                print(f"Copying from .xls file: {file_SOPO_path}...")
                copy_xls_to_sheet(file_SOPO_path, file_SOPO_path, template_wb, "3. Summary of Plan Ops")
                
            if os.path.exists(file_SOP_path):
                print(f"Copying from .xls file: {file_SOP_path}...")
                copy_xls_to_sheet(file_SOP_path, file_SOPO_path, template_wb, "4. Loan")
                
        elif file_extension == ".csv":
            # Copy from .csv file (if file exists)
            if os.path.exists(file_SNTA_path):
                print(f"Copying from .csv file: {file_SNTA_path}...")
                copy_csv_to_sheet(file_SNTA_path, file_SNTA_path, template_wb, "2. Summary of Net Trust Asset")
                # SOP0 file - Sheet TOTALS
            if os.path.exists(file_SOPO_path):
                print(f"Copying from .csv file: {file_SOPO_path}...")
                copy_csv_to_sheet(file_SOPO_path, file_SOPO_path, template_wb, "3. Summary of Plan Ops")
                
            if os.path.exists(file_SOP_path):
                print(f"Copying from .csv file: {file_SOP_path}...")
                copy_csv_to_sheet(file_SOP_path, file_SOPO_path, template_wb, "4. Loan")

        # results = summarize_hierarchical_df(
        #     df = totals_df,
        #     index_col = 0,
        #     sum_col = "TOTAL PLAN ACTIVITY",
        #     group_headers = ['Contributions/Employer', 'Contributions/Employee',
        #                      '"Adjustment (+)", "Adjustment (-)"', 'Administrative Fee', 'Realized Gain/(Loss)', 'Unrealized Gain/(Loss)', 'Interest and Dividends'
        #     ],
        #     roll_groups = ["Contributions/Employee"],
        #     combine_groups = {
        #         "net adj": ["Adjustment (+)", "Adjustment (-)"],  # Adjustment (+) and Adjustment (-) totals summed into one row
        #     },
        #     template_wb = template_wb,
        #     sheet_name = "6. SOPO Summary",
        # )

        # Code Generated by Sidekick is for learning and experimentation purposes only.
        results = summarize_hierarchical_df(
            df = totals_df,
            index_col = 0,
            sum_col = "TOTAL PLAN ACTIVITY",
            group_headers = [
                "Contributions/Employer", "Contributions/Employee",
                "Adjustment (+)", "Adjustment (-)",
                "Administrative Fee", "Realized Gain/(Loss)",
                "Unrealized Gain/(Loss)", "Interest and Dividends", "Benefit Payments"
            ],
            section_groups = [
                "Contributions/Employer", "Contributions/Employee",
                "Contributions/Employee",
            ],  # + only these have sub-rows
            roll_groups = [
                "Contributions/Employee",
            ],
            combine_groups = {
                "Adjustments": ["Adjustment (+)", "Adjustment (-)"],
                "Realized/Unrealized Gain Loss": ["Realized Gain/(Loss)", "Unrealized Gain/(Loss)"]
            },
            template_wb = template_wb,
            sheet_name = "6. SOPO Summary",
        )

        # Save the updated template
        print(f"Saving updated template to {output_path}...")
        template_wb.save(output_path)
        print("✓ Done! File saved successfully.")

    except FileNotFoundError as e:
        print(f"❌ Error: File not found - {e}")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()