# filename: python_backend/app.py
from flask import Flask, request, jsonify
import os
import openpyxl
from converter import convert_to_xlsx

app = Flask(__name__)

OUTPUT_DIR = os.environ.get(
    "CONVERTED_DIR",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "converted")
)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def excel_column_name(index):
    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/convert", methods=["POST"])
def convert():
    data = request.get_json() or {}
    file_paths = data.get("files", [])
    results = []

    for fp in file_paths:
        try:
            converted_path = convert_to_xlsx(fp, OUTPUT_DIR)
            results.append({
                "original": os.path.basename(fp),
                "converted": converted_path,
                "status": "ok"
            })
        except Exception as e:
            results.append({
                "original": os.path.basename(fp),
                "converted": None,
                "status": "error",
                "error": str(e)
            })

    return jsonify({"files": results})


@app.route("/sheets", methods=["POST"])
def get_sheets():
    data = request.get_json() or {}
    file_path = data.get("file")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return jsonify({"sheets": sheet_names})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/preview", methods=["POST"])
def preview():
    data = request.get_json() or {}
    file_path = data.get("file")
    sheet_name = data.get("sheet")
    max_rows = max(1, int(data.get("max_rows", 20)))

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True, max_row=max_rows))
        if not rows:
            wb.close()
            return jsonify({"columns": [], "rows": []})

        max_cols = 0
        for row in rows:
            for i in range(len(row) - 1, -1, -1):
                value = row[i]
                if value is not None and str(value).strip() != "":
                    max_cols = max(max_cols, i + 1)
                    break

        if max_cols == 0:
            wb.close()
            return jsonify({"columns": [], "rows": []})

        # Since converter.py permanently stripped empty columns, we can reliably sequence A, B, C...
        columns = [excel_column_name(i + 1) for i in range(max_cols)]

        preview_rows = []
        for row in rows:
            row_dict = {}
            for i in range(max_cols):
                row_dict[columns[i]] = row[i] if i < len(row) else None
            preview_rows.append(row_dict)

        wb.close()
        return jsonify({"columns": columns, "rows": preview_rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5123)