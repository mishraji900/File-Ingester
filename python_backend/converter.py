# filename: python_backend/converter.py
from __future__ import annotations

import os
import datetime
import xlrd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _sanitize(value: str) -> str:
    """strip special whitespace characters from string cell values."""
    return value.replace("\n", " ").replace("\t", " ").strip()


def _remove_empty_columns(ws) -> None:
    """Physically removes columns that contain absolutely no data to ensure contiguous A, B, C mapping."""
    if ws.max_column is None or ws.max_column == 0:
        return
        
    cols_to_delete = []
    # Traverse backwards so deleting a column doesn't shift the indices of columns we haven't checked yet
    for col_idx in range(ws.max_column, 0, -1):
        is_empty = True
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and str(val).strip() != "":
                is_empty = False
                break
        if is_empty:
            cols_to_delete.append(col_idx)

    for col_idx in cols_to_delete:
        ws.delete_cols(col_idx)


def convert_to_xlsx(file_path: str, output_dir: str) -> str:
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.xlsx")
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        _convert_modern(file_path, output_path)
    elif ext == ".xls":
        _convert_xls(file_path, output_path)
    else:
        raise ValueError(f"Unsupported file extension: {ext}")

    return output_path


def _convert_modern(file_path: str, output_path: str) -> None:
    wb = load_workbook(file_path, keep_vba=False)
    for ws in wb.worksheets:
        _remove_empty_columns(ws)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = _sanitize(cell.value)

    wb.save(output_path)


def _convert_xls(file_path: str, output_path: str) -> None:
    xls_wb = xlrd.open_workbook(file_path, formatting_info=True)
    xlsx_wb = openpyxl.Workbook()
    xlsx_wb.remove(xlsx_wb.active)

    for sheet_idx in range(xls_wb.nsheets):
        xls_ws = xls_wb.sheet_by_index(sheet_idx)
        xlsx_ws = xlsx_wb.create_sheet(title=xls_wb.sheet_names()[sheet_idx])

        # — Cell values + number formats ——————————————
        for row_idx in range(xls_ws.nrows):
            for col_idx in range(xls_ws.ncols):
                xls_cell = xls_ws.cell(row_idx, col_idx)
                xlsx_cell = xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1)
                xlsx_cell.value = _convert_cell_value(xls_cell, xls_wb)
                fmt = _get_format_string(xls_cell, xls_wb)
                if fmt:
                    xlsx_cell.number_format = fmt

        # — Merged cells ——————————————
        for rlo, rhi, clo, chi in xls_ws.merged_cells:
            xlsx_ws.merge_cells(
                start_row=rlo + 1,
                start_column=clo + 1,
                end_row=rhi,
                end_column=chi,
            )

        # — Column widths ——————————————
        for col_idx in range(xls_ws.ncols):
            col_info = xls_ws.colinfo_map.get(col_idx)
            if col_info and col_info.width > 0:
                col_letter = get_column_letter(col_idx + 1)
                xlsx_ws.column_dimensions[col_letter].width = col_info.width / 256.0
                
        # Apply empty column removal to the newly built xlsx sheet
        _remove_empty_columns(xlsx_ws)

    xlsx_wb.save(output_path)


def _convert_cell_value(cell: xlrd.sheet.Cell, wb: xlrd.Book):
    ctype, value = cell.ctype, cell.value

    if ctype == xlrd.XL_CELL_DATE:
        try:
            parts = xlrd.xldate_as_tuple(value, wb.datemode)
            return (
                datetime.time(parts[3], parts[4], int(parts[5]))
                if parts[0] == 0
                else datetime.datetime(*parts[:3], *parts[3:5], int(parts[5]))
            )
        except Exception:
            return value

    if ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(value)

    if ctype == xlrd.XL_CELL_ERROR:
        return None

    if isinstance(value, str):
        value = _sanitize(value)

    return value


def _get_format_string(cell: xlrd.sheet.Cell, wb: xlrd.Book) -> str | None:
    xf_idx = cell.xf_index
    if xf_idx is None:
        return None
    xf = wb.xf_list[xf_idx]
    fmt_str = wb.format_map[xf.format_key].format_str
    return fmt_str if fmt_str and fmt_str.upper() != "GENERAL" else None